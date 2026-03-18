import openpyxl
import json
import os

def parse_attendance(file_path):
    wb = openpyxl.load_workbook(file_path, data_only=True)
    attendance_data = []
    
    # We'll try to find all sheets that look like 'Week X'
    for sheet_name in wb.sheetnames:
        if 'Week' not in sheet_name:
            continue
            
        sheet = wb[sheet_name]
        week_num = sheet_name.replace('Week ', '').strip()
        
        # In the screenshot:
        # Row 4 has headers: Student, Email, and then Dates (23/02, 24/02, etc.)
        # Column A has names, starting from Row 5.
        
        dates = []
        for col in range(3, sheet.max_column + 1):
            date_val = sheet.cell(row=4, column=col).value
            if date_val:
                dates.append({
                    "col": col,
                    "date": str(date_val)
                })
        
        # Some sheets might have different layouts, but we'll stick to this pattern for now.
        for row_idx in range(5, sheet.max_row + 1):
            student_name = sheet.cell(row=row_idx, column=1).value
            if not student_name or student_name == "Student":
                continue
            
            # Use only first name as requested
            first_name = str(student_name).split()[0]
            
            for date_info in dates:
                cell = sheet.cell(row=row_idx, column=date_info["col"])
                # If there's an image, the underlying cell is usually empty (None).
                # The dashes are sometimes encoded as special characters.
                # So if the cell is completely empty, we assume the signature image is there (Present).
                # Otherwise, if it contains any text (like a dash), they are Absent.
                status = "present" if not cell.value else "absent"
                
                attendance_data.append({
                    "name": first_name,
                    "week": week_num,
                    "date": date_info["date"],
                    "status": status
                })
                
    return attendance_data

if __name__ == "__main__":
    path = r"c:/BUNYAN/Sem 8/Rules without Rulers/dataviz/attendance-Rules-w-o-Rulers---Thesis-Project (1).xlsx"
    data = parse_attendance(path)
    
    output_path = r"c:/BUNYAN/Sem 8/Rules without Rulers/dataviz/public/attendance_data.json"
    with open(output_path, 'w') as f:
        json.dump(data, f, indent=2)
    print(f"Data saved to {output_path}")
